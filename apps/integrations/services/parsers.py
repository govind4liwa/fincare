"""File parsing helpers — read CSV/XLSX into row dicts and coerce values.

Used by the bank-statement and platform-earnings importers. ``read_rows`` returns
``(headers, rows)`` where each row is a ``{header: value}`` dict, so the importer
maps headers to canonical fields via the ``ImportProfile.column_map``.
"""

import csv
import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

TWO_PLACES = Decimal("0.01")
ZERO = Decimal("0.00")
DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y")


class IntegrationError(ValueError):
    """Raised when a file cannot be parsed or fails validation."""


def file_hash(content):
    data = content.encode("utf-8") if isinstance(content, str) else content
    return hashlib.sha256(data).hexdigest()


def to_decimal(value):
    """Coerce a cell to a 2dp Decimal. Blank -> 0; handles commas and (negatives)."""
    if value is None or value == "":
        return ZERO
    if isinstance(value, int | float | Decimal):
        return Decimal(str(value)).quantize(TWO_PLACES)
    text = str(value).strip().replace(",", "").replace("AED", "").replace(" ", "")
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").strip()
    if text in ("", "-"):
        return ZERO
    try:
        amount = Decimal(text)
    except InvalidOperation as exc:
        raise IntegrationError(f"Cannot parse amount {value!r}.") from exc
    amount = amount.quantize(TWO_PLACES)
    return -amount if negative else amount


def to_date(value, fmt=""):
    """Coerce a cell to a ``date``. Accepts date/datetime objects or strings."""
    if value is None or value == "":
        raise IntegrationError("Missing date value.")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    formats = (fmt,) if fmt else DATE_FORMATS
    for f in formats:
        try:
            return datetime.strptime(text, f).date()
        except ValueError:
            continue
    raise IntegrationError(f"Cannot parse date {value!r} (formats tried: {', '.join(formats)}).")


def _rows_from_csv(content, skip_rows):
    text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
    reader = csv.reader(StringIO(text))
    table = [r for r in reader if any(str(c).strip() for c in r)]  # drop blank lines
    return table[skip_rows:]


def _rows_from_xlsx(content, skip_rows, sheet):
    from openpyxl import load_workbook

    data = content if isinstance(content, bytes) else content.encode()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    table = [list(r) for r in ws.iter_rows(values_only=True)]
    table = [r for r in table if any(c not in (None, "") for c in r)]
    return table[skip_rows:]


def read_rows(content, filename, *, skip_rows=0, sheet=""):
    """Return ``(headers, rows)`` for a CSV or XLSX file (by extension)."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        table = _rows_from_csv(content, skip_rows)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        table = _rows_from_xlsx(content, skip_rows, sheet)
    else:
        raise IntegrationError(f"Unsupported file type for {filename!r} (use .csv or .xlsx).")
    if not table:
        raise IntegrationError("File has no data rows.")

    headers = [str(h).strip() if h is not None else "" for h in table[0]]
    rows = []
    for raw in table[1:]:
        cells = list(raw) + [None] * (len(headers) - len(raw))
        rows.append(dict(zip(headers, cells, strict=False)))
    return headers, rows
