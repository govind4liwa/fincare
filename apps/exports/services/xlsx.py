"""Render a tabular report (catalog payload) to an Excel workbook (openpyxl)."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def report_to_xlsx_bytes(report):
    """Return .xlsx bytes for a ``{title, columns, rows}`` report payload."""
    wb = Workbook()
    ws = wb.active
    ws.title = report.get("code", "Report")[:31]

    bold = Font(bold=True)
    ws.append([report.get("title", "Report")])
    ws["A1"].font = bold

    ws.append(list(report["columns"]))
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = bold

    for row in report["rows"]:
        ws.append(list(row))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
